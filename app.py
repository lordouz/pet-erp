import streamlit as st
import pandas as pd
from datetime import datetime, date
import io

# 1. SAYFA VE TASARIM AYARLARI
st.set_page_config(page_title="PET Resin Komple ERP v2.6", layout="wide")

# MERKEZİ STOK KARTLARI VE BİRİMLERİ HAFIZASI
if 'stok_kartlari' not in st.session_state:
    st.session_state.stok_kartlari = {
        "Hammadde": [
            {"Ad": "PTA", "Birim": "Kg"}, {"Ad": "MEG", "Birim": "Kg"},
            {"Ad": "IPA", "Birim": "Kg"}, {"Ad": "DEG", "Birim": "Kg"}
        ],
        "Yardımcı Kimyasal": [
            {"Ad": "Antimon", "Birim": "Kg"}, {"Ad": "Fosforik Asit", "Birim": "Kg"},
            {"Ad": "Mavi Boya", "Birim": "Kg"}, {"Ad": "Kırmızı Boya", "Birim": "Kg"}
        ],
        "Ambalaj": [
            {"Ad": "PET Big Bag Çuval", "Birim": "Adet"}, {"Ad": "Ahşap Palet", "Birim": "Adet"}
        ],
        "Ara Mamul": [
            {"Ad": "Standart Amorf Chips", "Birim": "Kg"}
        ]
    }

# 2. MERKEZİ VERİ TABANI SİMÜLASYONU
if 'hammadde_depo' not in st.session_state:
    st.session_state.hammadde_depo = [
        {"Giriş Tarihi": "2026-08-15", "Depo": "Depo 1", "Kategori": "Hammadde", "Hammadde": "PTA", "LOT No": "PTA-LOT-001", "Miktar": 100000.0, "Birim": "Kg"},
        {"Giriş Tarihi": "2026-08-15", "Depo": "Depo 2", "Kategori": "Hammadde", "Hammadde": "MEG", "LOT No": "MEG-LOT-001", "Miktar": 50000.0, "Birim": "Kg"},
        {"Giriş Tarihi": "2026-08-15", "Depo": "Depo 1", "Kategori": "Yardımcı Kimyasal", "Hammadde": "Antimon", "LOT No": "ANT-LOT-001", "Miktar": 5000.0, "Birim": "Kg"},
        {"Giriş Tarihi": "2026-08-15", "Depo": "Depo 3", "Kategori": "Ambalaj", "Hammadde": "PET Big Bag Çuval", "LOT No": "BB-LOT-01", "Miktar": 500.0, "Birim": "Adet"},
        {"Giriş Tarihi": "2026-08-15", "Depo": "Depo 2", "Kategori": "Ara Mamul", "Hammadde": "Standart Amorf Chips", "LOT No": "AMF-LOT-00", "Miktar": 10000.0, "Birim": "Kg"}
    ]

if 'hammadde_kullanilan_toplam' not in st.session_state:
    st.session_state.hammadde_kullanilan_toplam = {}

if 'receteler' not in st.session_state:
    st.session_state.receteler = [
        {"Reçete Adı": "Standart Amorf Chips (Reaktör)", "Tür": "Ara Mamul Reçetesi", "BOM": {"PTA": 0.850000, "MEG": 0.135000, "Antimon": 0.005000}}
    ]

if 'mamul_depo' not in st.session_state:
    st.session_state.mamul_depo = []

if 'sevkiyat_depo' not in st.session_state:
    st.session_state.sevkiyat_depo = []

if 'uretim_harcamalari_log' not in st.session_state:
    st.session_state.uretim_harcamalari_log = []
def malzeme_birimi_bul(malzeme_adi):
    for kat, kalemler in st.session_state.stok_kartlari.items():
        for k in kalemler:
            if k["Ad"] == malzeme_adi: return k["Birim"]
    return "Kg"

def malzeme_depo_stok_getir(malzeme_adi, depo_adi):
    df = pd.DataFrame(st.session_state.hammadde_depo)
    if df.empty: return 0.0
    
    filtre = df[(df["Hammadde"] == malzeme_adi) & (df["Depo"] == depo_adi)]
    giren_toplam = filtre["Miktar"].sum() if not filtre.empty else 0.0
    
    anahtar = f"{malzeme_adi}_{depo_adi}"
    harcanan_toplam = st.session_state.hammadde_kullanilan_toplam.get(anahtar, 0.0)
    
    return max(0.0, giren_toplam - harcanan_toplam)

def endustriyel_excel_rapor_olustur(bas_tarih, bit_tarih):
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    buffer = io.BytesIO()
    
    df_depo = pd.DataFrame(st.session_state.hammadde_depo) if st.session_state.hammadde_depo else pd.DataFrame(columns=["Giriş Tarihi", "Depo", "Kategori", "Hammadde", "LOT No", "Miktar", "Birim"])
    df_harcama = pd.DataFrame(st.session_state.uretim_harcamalari_log) if st.session_state.uretim_harcamalari_log else pd.DataFrame(columns=["Tarih", "Harcanan Depo", "Üretim LOT", "Harcanan Malzeme", "Miktar", "Birim"])
    df_mamul = pd.DataFrame(st.session_state.mamul_depo) if st.session_state.mamul_depo else pd.DataFrame(columns=["Üretim Tarihi", "Ürün", "Üretim LOT / Silo", "Miktar"])
    df_sevk = pd.DataFrame(st.session_state.sevkiyat_depo) if st.session_state.sevkiyat_depo else pd.DataFrame(columns=["Sevkiyat Tarihi", "Müşteri", "İrsaliye No", "Plaka", "Ürün", "Sevk Edilen LOT", "Sevk Miktarı (Kg)"])
    
    def tarih_filtrele_ve_temizle(df, tarih_kolonu):
        if df.empty or tarih_kolonu not in df.columns: return df
        df_copy = df.copy()
        df_copy["temp_tarih"] = df_copy[tarih_kolonu].astype(str).apply(lambda x: x.split(" "))
        df_copy["temp_tarih"] = pd.to_datetime(df_copy["temp_tarih"], errors='coerce').dt.date
        filtered_df = df_copy[(df_copy["temp_tarih"] >= bas_tarih) & (df_copy["temp_tarih"] <= bit_tarih)]
        if "temp_tarih" in filtered_df.columns: filtered_df = filtered_df.drop(columns=["temp_tarih"])
        return filtered_df

    f_depo_giris = tarih_filtrele_ve_temizle(df_depo, "Giriş Tarihi")
    f_uretim_harcama = tarih_filtrele_ve_temizle(df_harcama, "Tarih")
    f_mamul_depo = tarih_filtrele_ve_temizle(df_mamul, "Üretim Tarihi")
    f_sevk_hareket = tarih_filtrele_ve_temizle(df_sevk, "Sevkiyat Tarihi")
    
    bakiye_satirlari = []
    tum_malzemeler = df_depo["Hammadde"].unique().tolist() if not df_depo.empty else []
    for m in tum_malzemeler:
        d1 = malzeme_depo_stok_getir(m, "Depo 1")
        d2 = malzeme_depo_stok_getir(m, "Depo 2")
        d3 = malzeme_depo_stok_getir(m, "Depo 3")
        bakiye_satirlari.append({
            "Malzeme / Ürün Adı": m, "Depo 1 Stok": d1, "Depo 2 Stok": d2, "Depo 3 Stok": d3,
            "Toplam Fabrika Stoğu": (d1 + d2 + d3), "Ölçü Birimi": malzeme_birimi_bul(m)
        })
    df_anlik_bakiye = pd.DataFrame(bakiye_satirlari) if bakiye_satirlari else pd.DataFrame(columns=["Malzeme / Ürün Adı", "Depo 1 Stok", "Depo 2 Stok", "Depo 3 Stok", "Toplam Fabrika Stoğu", "Ölçü Birimi"])

    with pd.ExcelWriter(buffer, engine='openpyxl') as writer:
        df_anlik_bakiye.to_excel(writer, index=False, sheet_name='Anlık Depolar Özet Matris')
        f_depo_giris.to_excel(writer, index=False, sheet_name='Giriş Hareketleri Detay')
        f_uretim_harcama.to_excel(writer, index=False, sheet_name='Üretim Tüketim Sarfiyat Detay')
        f_mamul_depo.to_excel(writer, index=False, sheet_name='Mamul Üretim Giriş Detay')
        f_sevk_hareket.to_excel(writer, index=False, sheet_name='Müşteri Sevkiyat İrsaliye Detay')
        
        header_font = Font(name='Segoe UI', size=11, bold=True, color='FFFFFF')
        header_fill = PatternFill(start_color='1F4E78', end_color='1F4E78', fill_type='solid')
        data_font = Font(name='Segoe UI', size=10)
        zebra_fill = PatternFill(start_color='F2F4F7', end_color='F2F4F7', fill_type='solid')
        thin_side = Side(border_style="thin", color="D9D9D9")
        thin_border = Border(left=thin_side, right=thin_side, top=thin_side, bottom=thin_side)
        
        for sheet_name in writer.sheets:
            ws = writer.sheets[sheet_name]
            ws.views.sheetView.showGridLines = True
            for cell in ws:
                cell.font = header_font; cell.fill = header_fill; cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            ws.row_dimensions.height = 28
            
            for row_idx in range(2, ws.max_row + 1):
                ws.row_dimensions[row_idx].height = 20
                for col_idx in range(1, ws.max_column + 1):
                    cell = ws.cell(row=row_idx, column=col_idx)
                    cell.font = data_font; cell.border = thin_border
                    if row_idx % 2 == 0: cell.fill = zebra_fill
                    
                    val = cell.value
                    col_name = str(ws.cell(row=1, column=col_idx).value)
                    if isinstance(val, (int, float)):
                        cell.alignment = Alignment(horizontal='right', vertical='center')
                        cell.number_format = '#,##0.000000' if "Oran" in col_name else '#,##0.0'
                    elif any(k in col_name for k in ["LOT", "No", "Plaka", "Tarih", "Depo"]):
                        cell.alignment = Alignment(horizontal='center', vertical='center')
                    else: cell.alignment = Alignment(horizontal='left', vertical='center')
                    
            for col in ws.columns:
                max_len = max(len(str(cell.value or '')) for cell in col)
                col_letter = get_column_letter(col.column)
                ws.column_dimensions[col_letter].width = max(max_len + 4, 14)
                
    return buffer.getvalue()
# 3. YAN PANEL MENÜ SİSTEMİ
st.sidebar.title("🧪 PET Resin ERP v2.6")
st.sidebar.write("---")
sayfa = st.sidebar.radio("Gitmek İstediğiniz Sayfa:", [
    "📊 Genel Depo & Stok Durumu",
    "📈 📊 Fabrika Raporlar Sayfası",
    "🗂️ 0. Stok Kartı Tanımlama Sayfası",
    "📥 1. Hammadde Giriş Sayfası",
    "📝 2. Reçete Oluşturma Sayfası",
    "🏭 3. Üretim Emri & Giriş Sayfası",
    "🚚 4. Müşteri Sevkiyat Sayfası"
])

# ==========================================
# SAYFA: GENEL DEPO VE STOK DURUMU
# ==========================================
if sayfa == "📊 Genel Depo & Stok Durumu":
    st.header("📊 Fabrika Çoklu Depo Yönetim Kokpiti")
    st.info("💡 Her malzemenin Depo 1, Depo 2 ve Depo 3 alanlarındaki anlık kırılımlarını görmek için ilgili sekmeye tıklayınız.")
    
    df_merkez = pd.DataFrame(st.session_state.hammadde_depo)
    kat_rehber = df_merkez.set_index("Hammadde")["Kategori"].to_dict() if not df_merkez.empty else {}
    tum_mevcut_malzemeler = df_merkez["Hammadde"].unique().tolist() if not df_merkez.empty else []

    # 1) Hammaddeler Başlığı
    with st.expander("🛠️ 1) HAMMADDE DEPOLARI DETAYLARI"):
        for h in tum_mevcut_malzemeler:
            if kat_rehber.get(h) == "Hammadde":
                st.markdown(f"**🔹 {h} Detay Matrisi ({malzeme_birimi_bul(h)}):**")
                c1, c2, c3, c4 = st.columns(4)
                d1 = malzeme_depo_stok_getir(h, "Depo 1")
                d2 = malzeme_depo_stok_getir(h, "Depo 2")
                d3 = malzeme_depo_stok_getir(h, "Depo 3")
                c1.metric("Depo 1", f"{d1:,.1f}")
                c2.metric("Depo 2", f"{d2:,.1f}")
                c3.metric("Depo 3", f"{d3:,.1f}")
                c4.metric("Küm. Toplam", f"{(d1+d2+d3):,.1f}")
                st.write(" ")
    
    # 2) Yardımcı Kimyasallar Başlığı
    with st.expander("🧪 2) YARDIMCI KİMYASAL DEPOLARI DETAYLARI"):
        for h in tum_mevcut_malzemeler:
            if kat_rehber.get(h) == "Yardımcı Kimyasal":
                st.markdown(f"**🔹 {h} Detay Matrisi ({malzeme_birimi_bul(h)}):**")
                c1, c2, c3, c4 = st.columns(4)
                d1 = malzeme_depo_stok_getir(h, "Depo 1")
                d2 = malzeme_depo_stok_getir(h, "Depo 2")
                d3 = malzeme_depo_stok_getir(h, "Depo 3")
                c1.metric("Depo 1", f"{d1:,.1f}")
                c2.metric("Depo 2", f"{d2:,.1f}")
                c3.metric("Depo 3", f"{d3:,.1f}")
                c4.metric("Küm. Toplam", f"{(d1+d2+d3):,.1f}")
                st.write(" ")

    # 3) Ambalaj Başlığı
    with st.expander("📦 3) AMBALAJ MALZEMESİ DEPOLARI DETAYLARI"):
        for h in tum_mevcut_malzemeler:
            if kat_rehber.get(h) == "Ambalaj":
                st.markdown(f"**🔹 {h} Detay Matrisi ({malzeme_birimi_bul(h)}):**")
                c1, c2, c3, c4 = st.columns(4)
                d1 = malzeme_depo_stok_getir(h, "Depo 1")
                d2 = malzeme_depo_stok_getir(h, "Depo 2")
                d3 = malzeme_depo_stok_getir(h, "Depo 3")
                c1.metric("Depo 1", f"{d1:,.0f}")
                c2.metric("Depo 2", f"{d2:,.0f}")
                c3.metric("Depo 3", f"{d3:,.0f}")
                c4.metric("Küm. Toplam", f"{(d1+d2+d3):,.0f}")
                st.write(" ")

    # 4) Ara Mamul Başlığı
    with st.expander("⚙️ 4) ARA MAMUL DEPOLARI DETAYLARI"):
        for h in tum_mevcut_malzemeler:
            if kat_rehber.get(h) == "Ara Mamul":
                st.markdown(f"**🔹 {h} Detay Matrisi ({malzeme_birimi_bul(h)}):**")
                c1, c2, c3, c4 = st.columns(4)
                d1 = malzeme_depo_stok_getir(h, "Depo 1")
                d2 = malzeme_depo_stok_getir(h, "Depo 2")
                d3 = malzeme_depo_stok_getir(h, "Depo 3")
                c1.metric("Depo 1", f"{d1:,.1f}")
                c2.metric("Depo 2", f"{d2:,.1f}")
                c3.metric("Depo 3", f"{d3:,.1f}")
                c4.metric("Küm. Toplam", f"{(d1+d2+d3):,.1f}")
                st.write(" ")

    # 5) Ürün Bazlı Gruplanmış Satışa Hazır Mamul Başlığı (YAZIM HATASI GİDERİLDİ)
    with st.expander("🏭 5) ÜRÜN BAZLI SATIŞA HAZIR MAMUL DEPOSU"):
        if st.session_state.mamul_depo:
            df_mamul = pd.DataFrame(st.session_state.mamul_depo)
            df_sevk_matris = pd.DataFrame(st.session_state.sevkiyat_depo) if st.session_state.sevkiyat_depo else pd.DataFrame(columns=["Ürün", "Sevk Edilen LOT", "Sevk Miktarı (Kg)"])
            for urun_adi, urun_data in df_mamul.groupby("Ürün"):
                toplam_sevk_edilen = df_sevk_matris[df_sevk_matris["Ürün"] == urun_adi]["Sevk Miktarı (Kg)"].sum() if not df_sevk_matris.empty else 0.0
                toplam_uretim = urun_data["Miktar"].sum()
                kalan_net_mamul = max(0.0, toplam_uretim - toplam_sevk_edilen)
                st.write(f"**🔹 {urun_adi}** | Kalan Net Satış Stoğu: **{kalan_net_mamul:,.1f} Kg**")
                lot_satirlari = []
                for lot_no, lot_data in urun_data.groupby("Üretim LOT / Silo"):
                    l_uretim = lot_data["Miktar"].sum()
                    # HATA DÜZELTİLDİ: 'lot_no balance' yazım hatası temizlendi
                    l_sevk = df_sevk_matris[(df_sevk_matris["Ürün"] == urun_adi) & (df_sevk_matris["Sevk Edilen LOT"] == lot_no)]["Sevk Miktarı (Kg)"].sum() if not df_sevk_matris.empty else 0.0
                    lot_satirlari.append({"Üretim LOT / Silo": lot_no, "Üretilen (Kg)": l_uretim, "Sevk Edilen (Kg)": l_sevk, "Kalan Stok (Kg)": max(0.0, l_uretim - l_sevk)})
                st.dataframe(pd.DataFrame(lot_satirlari), use_container_width=True)
        else: st.info("Satışa hazır bitmiş mamul stoku bulunmuyor.")

# ==========================================
# SAYFA: FABRİKA RAPORLAR SAYFASI
# ==========================================
elif sayfa == "📈 📊 Fabrika Raporlar Sayfası":
    st.header("📈 📊 Dönemsel Fabrika Profesyonel Rapor İstasyonu")
    c_t1, c_t2 = st.columns(2)
    bas_secim = c_t1.date_input("Analiz Başlangıç Tarihi", value=date(2026, 1, 1))
    bit_secim = c_t2.date_input("Analiz Bitiş Tarihi", value=date(2026, 12, 31))
    if bas_secim <= bit_secim:
        excel_dosyası = endustriyel_excel_rapor_olustur(bas_secim, bit_secim)
        st.download_button(label="📊 Fabrika Konsolide Profesyonel Raporu İndir (.XLSX)", data=excel_dosyası, file_name=f"Fabrika_Kurumsal_Rapor_{bas_secim}_to_{bit_secim}.xlsx", mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
# ==========================================
# SAYFA 0: BİRİMLİ STOK KARTI TANIMLAMA
# ==========================================
elif sayfa == "🗂️ 0. Stok Kartı Tanımlama Sayfası":
    st.header("🗂️ Fabrika Malzeme / Stok Kartı Tanımlama")
    with st.form("stok_kart_form"):
        k_kat = st.selectbox("Kartın Bağlanacağı Kategori", ["Hammadde", "Yardımcı Kimyasal", "Ambalaj", "Ara Mamul"])
        k_adi = st.text_input("Yeni Malzeme / Stok Kartı Adı")
        k_birim = st.selectbox("Miktar Ölçü Birimi", ["Kg", "Adet", "Ton"])
        if st.form_submit_button("Yeni Stok Kartını Sisteme Kaydet") and k_adi:
            st.session_state.stok_kartlari[k_kat].append({"Ad": k_adi, "Birim": k_birim})
            st.success(f"✅ '{k_adi}' kartı kaydedildi."); st.rerun()
            
    st.subheader("📋 Sistemde Kayıtlı Kart Listesi")
    for k, v in st.session_state.stok_kartlari.items():
        st.write(f"**{k}:** {', '.join([f'{x[\"Ad\"]} ({x[\"Birim\"]})' for x in v])}")

# ==========================================
# SAYFA 1: HAMMADDE / MALZEME GİRİŞİ
# ==========================================
elif sayfa == "📥 1. Hammadde Giriş Sayfası":
    st.header("📥 Fabrika Depolarına Giriş Kabul Ekranı")
    kat_turu = st.selectbox("Malzeme Kategorisi Seçin", ["Hammadde", "Yardımcı Kimyasal", "Ambalaj", "Ara Mamul"])
    uygun_malzemeler = [x["Ad"] for x in st.session_state.stok_kartlari.get(kat_turu, [])]
    
    if uygun_malzemeler:
        with st.form("hammadde_form"):
            g_tarih = st.date_input("Giriş Tarihi", value=datetime.now())
            hedef_depo_secim = st.selectbox("Malzemenin İndirileceği Fiziksel Depo", ["Depo 1", "Depo 2", "Depo 3"])
            h_turu = st.selectbox("Malzeme / Kalem Adı", uygun_malzemeler)
            h_lot = st.text_input("Gelen LOT / Parti Numarası")
            secilen_birim = malzeme_birimi_bul(h_turu)
            h_miktar = st.number_input(f"Gelen Miktar ({secilen_birim})", min_value=0.1, step=50.0)
            
            if st.form_submit_button("Malzemeyi Depoya Kabul Et") and h_lot:
                st.session_state.hammadde_depo.append({
                    "Giriş Tarihi": str(g_tarih), "Depo": hedef_depo_secim, "Kategori": kat_turu, 
                    "Hammadde": h_turu, "LOT No": h_lot, "Miktar": h_miktar, "Birim": secilen_birim
                })
                st.success(f"✅ {h_turu} başarıyla {hedef_depo_secim} alanına kabul edildi."); st.rerun()

# ==========================================
# SAYFA 2: REÇETE OLUŞTURMA VE DÜZENLEME
# ==========================================
elif sayfa == "📝 2. Reçete Oluşturma Sayfası":
    st.header("📝 Ürün Reçetesi (BOM) Yönetim İstasyonu")
    operasyon_turu = st.radio("Yapmak İstediğiniz İşlem:", ["➕ Yeni Reçete Oluştur", "✏️ Mevcut Reçeteyi Gör ve Düzenle"])
    duzenlenecek_recete_adi, eski_bom, duzenleme_indeksi, r_adi_val, r_turu_idx = None, {}, None, "", 0
    
    if operasyon_turu == "✏️ Mevcut Reçeteyi Gör ve Düzenle" and st.session_state.receteler:
        recete_isimleri = [r["Reçete Adı"] for r in st.session_state.receteler]
        duzenlenecek_recete_adi = st.selectbox("🔍 Düzenlemek İstediğiniz Reçeteyi Seçin", recete_isimleri)
        duzenleme_indeksi = next(idx for idx, r in enumerate(st.session_state.receteler) if r["Reçete Adı"] == duzenlenecek_recete_adi)
        eski_bom, r_adi_val = st.session_state.receteler[duzenleme_indeksi]["BOM"], st.session_state.receteler[duzenleme_indeksi]["Reçete Adı"]
        r_turu_idx = 0 if st.session_state.receteler[duzenleme_indeksi]["Tür"] == "Ara Mamul Reçetesi" else 1
        st.subheader("📋 Kayıtlı Mevcut Reçete Oranları Özet Görünümü")
        if eski_bom:
            st.dataframe(pd.DataFrame([{"Malzeme Adı": m, "Katsayı Oranı": f"{o:.6f}", "Birim": malzeme_birimi_bul(m)} for m, o in eski_bom.items()]), use_container_width=True)

    r_adi_input = st.text_input("Reçete / Ürün Adı", value=r_adi_val)
    r_turu_input = st.selectbox("Reçete Sınıfı", ["Ara Mamul Reçetesi", "Mamul Reçetesi"], index=r_turu_idx)
    tab_ham, tab_kim, tab_amb, tab_ara = st.tabs(["🛠️ Hammaddeler", "🧪 Yardımcı Kimyasallar", "📦 Ambalaj", "⚙️ Ara Mamuller"])
    secilen_bom = {}
    
    with tab_ham:
        for k in st.session_state.stok_kartlari["Hammadde"]:
            val = st.number_input(f"{k['Ad']} ({k['Birim']})", min_value=0.0, max_value=100.0, value=float(eski_bom.get(k['Ad'], 0.0)), step=0.000001, format="%.6f", key=f"h_{k['Ad']}")
            if val > 0: secilen_bom[k['Ad']] = val
    with tab_kim:
        for k in st.session_state.stok_kartlari["Yardımcı Kimyasal"]:
            val = st.number_input(f"{k['Ad']} ({k['Birim']})", min_value=0.0, max_value=100.0, value=float(eski_bom.get(k['Ad'], 0.0)), step=0.000001, format="%.6f", key=f"ki_{k['Ad']}")
            if val > 0: secilen_bom[k['Ad']] = val
    with tab_amb:
        for k in st.session_state.stok_kartlari["Ambalaj"]:
            val = st.number_input(f"{k['Ad']} ({k['Birim']})", min_value=0.0, max_value=100.0, value=float(eski_bom.get(k['Ad'], 0.0)), step=0.000001, format="%.6f", key=f"am_{k['Ad']}")
            if val > 0: secilen_bom[k['Ad']] = val
    with tab_ara:
        df_d = pd.DataFrame(st.session_state.hammadde_depo)
        stoktaki_ara = df_d[df_d["Kategori"] == "Ara Mamul"]["Hammadde"].unique().tolist() if not df_d.empty else []
        for am_ad in list(set([x["Ad"] for x in st.session_state.stok_kartlari["Ara Mamul"]] + stoktaki_ara)):
            val = st.number_input(f"{am_ad} (Kg)", min_value=0.0, max_value=100.0, value=float(eski_bom.get(am_ad, 0.0)), step=0.000001, format="%.6f", key=f"ar_{am_ad}")
            if val > 0: secilen_bom[am_ad] = val

    if st.button("💾 Değişiklikleri Onayla") and r_adi_input and secilen_bom:
        g_rec = {"Reçete Adı": r_adi_input, "Tür": r_turu_input, "BOM": secilen_bom}
        if operasyon_turu == "✏️ Mevcut Reçeteyi Gör ve Düzenle" and duzenleme_indeksi is not None: st.session_state.receteler[duzenleme_indeksi] = g_rec
        else: st.session_state.receteler.append(g_rec)
        st.success("✅ Reçete Kaydedildi!"); st.rerun()
# ==========================================
# SAYFA 3: ÜRETİM EMRİ & GİRİŞİ
# ==========================================
elif sayfa == "🏭 3. Üretim Emri & Giriş Sayfası":
    st.header("🏭 Üretim Yönetim ve Reaktör Besleme Arayüzü")
    u_kategori = st.radio("Yapılacak Üretim Sınıfı:", ["1) Ara Mamul Üretimi", "2) Mamul Üretimi"])
    hedef_tur = "Ara Mamul Reçetesi" if "1)" in u_kategori else "Mamul Reçetesi"
    uygun_receteler = [r for r in st.session_state.receteler if r.get("Tür") == hedef_tur]
    
    if uygun_receteler:
        secilen_recete_adi = st.selectbox("Kullanılacak Reçeteyi Seçin", [r["Reçete Adı"] for r in uygun_receteler])
        hedef_miktar = st.number_input("Hedef Üretim Hacmi (Kg)", min_value=1.0, value=1000.0)
        u_lot = st.text_input("Üretim Parti / Silo LOT No", value=f"LOT-{datetime.now().strftime('%Y%m%d%H%M')}")
        kaynak_depo_secim = st.selectbox("🚀 Hammaddelerin Tüketileceği Kaynak Depo Alanı", ["Depo 1", "Depo 2", "Depo 3"])
        
        secilen_recete = next(r for r in uygun_receteler if r["Reçete Adı"] == secilen_recete_adi)
        
        with st.form("uretim_form"):
            fiili_girisler = {}
            for h_adi, oran in secilen_recete["BOM"].items():
                teorik = hedef_miktar * oran
                m_stok = malzeme_depo_stok_getir(h_adi, kaynak_depo_secim)
                fiili_girisler[h_adi] = st.number_input(f"{h_adi} Fiili Tüketim ({malzeme_birimi_bul(h_adi)}) [Mevcut {kaynak_depo_secim}: {m_stok:,.1f}]", min_value=0.0, value=float(teorik))
                
            if st.form_submit_button("Üretimi Onayla ve Depodan Tüket"):
                kontrol = True
                eksikler = []
                for h_adi, f_amt in fiili_girisler.items():
                    m_stok = malzeme_depo_stok_getir(h_adi, kaynak_depo_secim)
                    if m_stok < f_amt:
                        kontrol = False
                        eksikler.append(f"{h_adi} ({kaynak_depo_secim} Yetersiz)")
                        
                if not kontrol: 
                    st.error(f"❌ Üretim Reddedildi! Seçilen {kaynak_depo_secim} alanında yetersiz bakiye var: {', '.join(eksikler)}")
                else:
                    current_date_str = datetime.now().strftime("%Y-%m-%d %H:%M")
                    for h_adi, f_amt in fiili_girisler.items():
                        anahtar = f"{h_adi}_{kaynak_depo_secim}"
                        st.session_state.hammadde_kullanilan_toplam[anahtar] = st.session_state.hammadde_kullanilan_toplam.get(anahtar, 0.0) + f_amt
                        st.session_state.uretim_harcamalari_log.append({
                            "Tarih": current_date_str, "Harcanan Depo": kaynak_depo_secim, "Üretim LOT": u_lot, 
                            "Harcanan Malzeme": h_adi, "Miktar": f_amt, "Birim": malzeme_birimi_bul(h_adi)
                        })
                    
                    if hedef_tur == "Ara Mamul Reçetesi":
                        st.session_state.hammadde_depo.append({"Giriş Tarihi": datetime.now().strftime("%Y-%m-%d"), "Depo": kaynak_depo_secim, "Kategori": "Ara Mamul", "Hammadde": secilen_recete_adi, "LOT No": u_lot, "Miktar": hedef_miktar, "Birim": "Kg"})
                    else:
                        st.session_state.mamul_depo.append({"Üretim Tarihi": current_date_str, "Ürün": secilen_recete_adi, "Üretim LOT / Silo": u_lot, "Miktar": hedef_miktar})
                    st.success(f"🎉 Üretim tamamlandı! Hammaddeler başarıyla {kaynak_depo_secim} stoklarından düşüldü."); st.rerun()

# ==========================================
# SAYFA 4: MÜŞTERİ SEVKİYAT VE İRSALİYE ÇIKIŞ SAYFASI
# ==========================================
elif sayfa == "🚚 4. Müşteri Sevkiyat Sayfası":
    st.header("🚚 Müşteri Mamul Sevkiyat ve İrsaliye İstasyonu")
    
    if not st.session_state.mamul_depo:
        st.warning("⚠️ Sevkiyat yapabilmek için mamul deposunda ürün bulunmalıdır.")
    else:
        df_mamul = pd.DataFrame(st.session_state.mamul_depo)
        df_sevk_matris = pd.DataFrame(st.session_state.sevkiyat_depo) if st.session_state.sevkiyat_depo else pd.DataFrame(columns=["Ürün", "Sevk Edilen LOT", "Sevk Miktarı (Kg)"])
        
        aktif_urunler = df_mamul["Ürün"].unique().tolist()
        secilen_sevk_urun = st.selectbox("Sevk Edilecek Mamul Ürünü Seçin", aktif_urunler)
        
        uygun_lotlar = df_mamul[df_mamul["Ürün"] == secilen_sevk_urun]["Üretim LOT / Silo"].unique().tolist()
        secilen_sevk_lot = st.selectbox("Sevk Edilecek Üretim LOT / Silo Seçin", uygun_lotlar)
        
        l_toplam_uretim = df_mamul[(df_mamul["Ürün"] == secilen_sevk_urun) & (df_mamul["Üretim LOT / Silo"] == secilen_sevk_lot)]["Miktar"].sum()
        l_toplam_sevk = df_sevk_matris[(df_sevk_matris["Ürün"] == secilen_sevk_urun) & (df_sevk_matris["Sevk Edilen LOT"] == secilen_sevk_lot)]["Sevk Miktarı (Kg)"].sum() if not df_sevk_matris.empty else 0.0
        
        mevcut_lot_bakiyesi = max(0.0, l_toplam_uretim - l_toplam_sevk)
        st.info(f"💡 Seçilen {secilen_sevk_lot} lot numaralı ürünün kullanılabilir net bakiyesi: **{mevcut_lot_bakiyesi:,.1f} Kg**")
        
        with st.form("sevkiyat_form"):
            c_s1, c_s2 = st.columns(2)
            f_musteri = c_s1.text_input("Müşteri Firma Adı")
            f_irsaliye = c_s2.text_input("İrsaliye / Fatura Numarası")
            
            c_s3, c_s4 = st.columns(2)
            f_plaka = c_s3.text_input("Nakliye Araç Plakası")
            f_sevk_miktar = c_s4.number_input("Sevk Edilecek Net Miktar (Kg)", min_value=1.0, value=float(min(1000.0, mevcut_lot_bakiyesi)), step=100.0)
            
            if st.form_submit_button("Sevkiyatı Onayla ve İrsaliye Kes"):
                if not f_musteri or not f_irsaliye: st.error("❌ Boş bırakılamaz!")
                elif f_sevk_miktar > mevcut_lot_bakiyesi: st.error("❌ Stok Yetersiz!")
                else:
                    st.session_state.sevkiyat_depo.append({
                        "Sevkiyat Tarihi": datetime.now().strftime("%Y-%m-%d %H:%M"), "Müşteri": f_musteri, "İrsaliye No": f_irsaliye,
                        "Plaka": f_plaka, "Ürün": secilen_sevk_urun, "Sevk Edilen LOT": secilen_sevk_lot, "Sevk Miktarı (Kg)": f_sevk_miktar
                    })
                    st.success("🎉 Sevkiyat tamamlandı!"); st.rerun()

        st.write("---")
        st.subheader("📋 Geçmiş Müşteri Sevkiyat ve İrsaliye Kayıt Listesi")
        if st.session_state.sevkiyat_depo: st.dataframe(pd.DataFrame(st.session_state.sevkiyat_depo), use_container_width=True)
